"""Builds the monthly accounting report as a structured Excel (.xlsx) workbook.

Pure rendering: it takes the summary already computed by AccountingService and
produces .xlsx bytes — no DB access, easy to test. Two sheets:
  - "Resumen": per-client totals (expected / paid / previous credit / balance).
  - "Detalle por contrato": per-contract class breakdown.
"""
import io

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.worksheet.worksheet import Worksheet

from app.schemas.accounting import AccountingSummaryEntryResponse
from app.services.accounting_service import MONTH_NAMES

EURO_FMT = '#,##0.00 €'
TITLE_FONT = Font(bold=True, size=14)
HEADER_FONT = Font(bold=True)
DEBT_FONT = Font(color="C00000")
CREDIT_FONT = Font(color="0070C0")


def build_monthly_report_xlsx(
    summary: list[AccountingSummaryEntryResponse], month: int, year: int
) -> bytes:
    """Render the monthly accounting summary into .xlsx bytes."""
    period = f"{MONTH_NAMES[month - 1]} {year}"

    wb = Workbook()
    _build_resumen(wb.active, summary, period)
    _build_detalle(wb.create_sheet("Detalle por contrato"), summary, period)

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _write_headers(ws: Worksheet, headers: list[str], row: int) -> None:
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=text)
        cell.font = HEADER_FONT


def _build_resumen(
    ws: Worksheet, summary: list[AccountingSummaryEntryResponse], period: str
) -> None:
    ws.title = "Resumen"
    ws["A1"] = f"Contabilidad — {period}"
    ws["A1"].font = TITLE_FONT

    _write_headers(ws, ["Cliente", "Esperado", "Pagado", "Crédito previo", "Balance"], row=3)

    total_expected = total_paid = total_balance = 0.0
    row = 4
    for entry in summary:
        ws.cell(row=row, column=1, value=entry.client_name)
        ws.cell(row=row, column=2, value=entry.expected).number_format = EURO_FMT
        ws.cell(row=row, column=3, value=entry.paid).number_format = EURO_FMT
        ws.cell(row=row, column=4, value=entry.previous_credit).number_format = EURO_FMT
        balance_cell = ws.cell(row=row, column=5, value=entry.balance)
        balance_cell.number_format = EURO_FMT
        if entry.balance < 0:
            balance_cell.font = DEBT_FONT
        elif entry.balance > 0:
            balance_cell.font = CREDIT_FONT

        total_expected += entry.expected
        total_paid += entry.paid
        total_balance += entry.balance
        row += 1

    ws.cell(row=row, column=1, value="Totales")
    ws.cell(row=row, column=2, value=round(total_expected, 2)).number_format = EURO_FMT
    ws.cell(row=row, column=3, value=round(total_paid, 2)).number_format = EURO_FMT
    ws.cell(row=row, column=5, value=round(total_balance, 2)).number_format = EURO_FMT
    for col in range(1, 6):
        ws.cell(row=row, column=col).font = HEADER_FONT

    for col, width in zip("ABCDE", (28, 14, 14, 16, 14)):
        ws.column_dimensions[col].width = width


def _build_detalle(
    ws: Worksheet, summary: list[AccountingSummaryEntryResponse], period: str
) -> None:
    ws["A1"] = f"Detalle por contrato — {period}"
    ws["A1"].font = TITLE_FONT

    _write_headers(
        ws,
        ["Cliente", "Contrato", "€/h", "Normales", "Canc. con pago", "Canc. sin pago", "Esperado"],
        row=3,
    )

    row = 4
    for entry in summary:
        for contract in entry.contracts:
            ws.cell(row=row, column=1, value=entry.client_name)
            ws.cell(row=row, column=2, value=contract.contract_description)
            ws.cell(row=row, column=3, value=contract.hourly_rate)
            ws.cell(row=row, column=4, value=contract.normal_count)
            ws.cell(row=row, column=5, value=contract.cancelled_with_payment_count)
            ws.cell(row=row, column=6, value=contract.cancelled_without_payment_count)
            ws.cell(row=row, column=7, value=contract.expected).number_format = EURO_FMT
            row += 1

    for col, width in zip("ABCDEFG", (28, 30, 8, 10, 16, 16, 14)):
        ws.column_dimensions[col].width = width
