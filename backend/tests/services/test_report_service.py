"""Tests for report_service — monthly accounting Excel (.xlsx) generation."""
import io

from openpyxl import load_workbook

from app.services.report_service import build_monthly_report_xlsx

SUMMARY = [
    {
        "client_id": 1, "client_name": "Ana García",
        "expected": 200.0, "paid": 150.0, "previous_credit": 0.0, "balance": -50.0,
        "month": 4, "year": 2026, "month_name": "Abril",
        "contracts": [
            {
                "contract_id": 10, "contract_description": "Clases de inglés",
                "hourly_rate": 25.0, "normal_count": 8,
                "cancelled_with_payment_count": 0, "cancelled_without_payment_count": 1,
                "expected": 200.0, "class_count": 8,
            },
        ],
    },
    {
        "client_id": 2, "client_name": "Beto López",
        "expected": 100.0, "paid": 120.0, "previous_credit": 0.0, "balance": 20.0,
        "month": 4, "year": 2026, "month_name": "Abril",
        "contracts": [],
    },
]


def _load(summary, month=4, year=2026):
    data = build_monthly_report_xlsx(summary, month, year)
    assert isinstance(data, (bytes, bytearray))
    return load_workbook(io.BytesIO(data))


class TestBuildMonthlyReport:
    def test_has_resumen_and_detalle_sheets(self):
        wb = _load(SUMMARY)
        assert wb.sheetnames == ["Resumen", "Detalle por contrato"]

    def test_resumen_title_includes_period(self):
        ws = _load(SUMMARY)["Resumen"]
        assert ws["A1"].value == "Contabilidad — Abril 2026"

    def test_resumen_has_per_client_rows(self):
        ws = _load(SUMMARY)["Resumen"]
        assert ws["A3"].value == "Cliente"
        assert ws["A4"].value == "Ana García"      # input order preserved
        assert ws["B4"].value == 200.0
        assert ws["C4"].value == 150.0
        assert ws["E4"].value == -50.0

    def test_resumen_totals_row(self):
        ws = _load(SUMMARY)["Resumen"]
        # two clients -> rows 4,5 -> totals at row 6
        assert ws["A6"].value == "Totales"
        assert ws["B6"].value == 300.0
        assert ws["C6"].value == 270.0
        assert ws["E6"].value == -30.0

    def test_detalle_has_contract_rows(self):
        ws = _load(SUMMARY)["Detalle por contrato"]
        assert ws["A3"].value == "Cliente"
        assert ws["A4"].value == "Ana García"      # only client with a contract
        assert ws["B4"].value == "Clases de inglés"
        assert ws["G4"].value == 200.0             # expected (last column)

    def test_empty_summary_still_builds_both_sheets(self):
        wb = _load([])
        assert wb.sheetnames == ["Resumen", "Detalle por contrato"]
