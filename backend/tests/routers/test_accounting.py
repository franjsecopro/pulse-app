"""Router tests for /api/accounting/report — the Excel report download."""
import io

from openpyxl import load_workbook
from httpx import AsyncClient

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


class TestMonthlyReport:
    async def test_returns_xlsx_with_attachment_filename(self, app_client: AsyncClient):
        response = await app_client.get("/api/accounting/report?month=4&year=2026")

        assert response.status_code == 200
        assert response.headers["content-type"] == XLSX_MEDIA_TYPE
        assert "accounting_abril_2026.xlsx" in response.headers["content-disposition"]

    async def test_body_is_a_valid_workbook(self, app_client: AsyncClient):
        response = await app_client.get("/api/accounting/report?month=4&year=2026")

        wb = load_workbook(io.BytesIO(response.content))
        assert wb.sheetnames == ["Resumen", "Detalle por contrato"]

    async def test_rejects_invalid_month(self, app_client: AsyncClient):
        response = await app_client.get("/api/accounting/report?month=13&year=2026")

        assert response.status_code == 422
